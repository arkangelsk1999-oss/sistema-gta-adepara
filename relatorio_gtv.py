"""
relatorio_gtv.py
Geração de Excel, CSV e LAI para o módulo GTV.
"""

import io
from datetime import datetime

# ─────────────────────────────────────────────
# CAMPOS PARA EXPORTAÇÃO NORMAL
# ─────────────────────────────────────────────

CAMPOS_EXIBICAO_GTV = [
    ("gtv_numero",        "Nº GTV"),
    ("data_emissao",      "Data Emissão"),
    ("situacao_pedido",   "Situação"),
    ("cultura",           "Cultura"),
    ("quantidade_carga",  "Quantidade"),
    ("medida",            "Medida"),
    ("valor",             "Valor R$"),
    ("procedencia",       "Procedência (Origem)"),
    ("municipio_origem",  "Município Origem"),
    ("destinatario_nome", "Destinatário"),
    ("municipio_destino", "Município Destino"),
    ("codigo_up_origem",  "Cód. UP Origem"),
    ("emitida_por_nome",  "Emitido Por"),
    ("veiculo",           "Veículo"),
    ("transito",          "Trânsito"),
    ("quantidade_chegada","Qtd. Chegada"),
]

# ─────────────────────────────────────────────
# CAMPOS LAI
# ─────────────────────────────────────────────

CAMPOS_LAI_GTV = [
    ("gtv_numero",        "Número da GTV"),
    ("data_emissao",      "Data de Emissão"),
    ("situacao_pedido",   "Situação"),
    ("cultura",           "Cultura"),
    ("quantidade_carga",  "Quantidade da Carga"),
    ("medida",            "Medida"),
    ("valor",             "Valor R$"),
    ("procedencia",       "Procedência (Origem)"),
    ("municipio_origem",  "Município de Origem"),
    ("destinatario_nome", "Destinatário"),
    ("municipio_destino", "Município de Destino"),
    ("codigo_up_origem",  "Código UP de Origem"),
    ("emitida_por_nome",  "Emissor"),
    ("veiculo",           "Veículo"),
    ("transito",          "Trânsito"),
    ("quantidade_chegada","Quantidade de Chegada"),
]


# ─────────────────────────────────────────────
# EXCEL — BUSCA NORMAL
# ─────────────────────────────────────────────

def gerar_excel_gtv(resultado, termo_busca, usuario):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise ImportError("openpyxl não instalado.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    azul_escuro = PatternFill("solid", fgColor="1A5276")
    azul_claro  = PatternFill("solid", fgColor="D6EAF8")
    branco_font = Font(color="FFFFFF", bold=True, size=11)

    dados_origem  = resultado.get('gtv', {}).get('origem', [])
    dados_destino = resultado.get('gtv', {}).get('destino', [])

    for titulo_aba, dados in [("Como Origem", dados_origem), ("Como Destino", dados_destino)]:
        if not dados:
            continue
        ws = wb.create_sheet(titulo_aba)

        cabecalhos = [label for _, label in CAMPOS_EXIBICAO_GTV]
        for col_idx, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_idx, value=cab)
            cell.fill = azul_escuro
            cell.font = branco_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, reg in enumerate(dados, 2):
            fill = azul_claro if row_idx % 2 == 0 else PatternFill()
            for col_idx, (campo, _) in enumerate(CAMPOS_EXIBICAO_GTV, 1):
                val = reg.get(campo, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
                cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Aba rastreio oculta
    ws_r = wb.create_sheet("_rastreio")
    ws_r.sheet_state = 'hidden'
    cpf_usuario = usuario.get('cpf', '00000000000')
    ws_r['A1'] = "Sistema GTA ADEPARÁ — Módulo GTV"
    ws_r['A2'] = f"Gerado por: {usuario.get('nome', '')}"
    ws_r['A3'] = f"E-mail: {usuario.get('email', '')}"
    ws_r['A4'] = f"Órgão: {usuario.get('orgao', '')}"
    ws_r['A5'] = f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_r['A6'] = f"Termo buscado: {termo_busca}"
    ws_r.protection.sheet = True
    ws_r.protection.password = cpf_usuario

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# CSV — BUSCA NORMAL
# ─────────────────────────────────────────────

def gerar_csv_gtv(resultado, termo_busca, usuario):
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')

    cabecalhos = [label for _, label in CAMPOS_EXIBICAO_GTV]
    writer.writerow(cabecalhos)

    for tipo in ['origem', 'destino']:
        for reg in resultado.get('gtv', {}).get(tipo, []):
            linha = [reg.get(campo, '') or '' for campo, _ in CAMPOS_EXIBICAO_GTV]
            writer.writerow(linha)

    buf.seek(0)
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


# ─────────────────────────────────────────────
# LAI GTV — Excel com filtros
# ─────────────────────────────────────────────

def gerar_excel_lai_gtv(registros, usuario):
    try:
        import xlsxwriter
        import zipfile
    except ImportError:
        raise ImportError("xlsxwriter não instalado.")

    buf_xlsx = io.BytesIO()
    wb = xlsxwriter.Workbook(buf_xlsx, {'in_memory': True})

    fmt_titulo  = wb.add_format({'bold': True, 'font_size': 13, 'font_color': '#1A5276'})
    fmt_cab     = wb.add_format({'bold': True, 'bg_color': '#1A5276', 'font_color': '#FFFFFF',
                                  'border': 1, 'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
    fmt_dado    = wb.add_format({'border': 1})
    fmt_dado_alt= wb.add_format({'border': 1, 'bg_color': '#EBF5FB'})
    fmt_aviso   = wb.add_format({'bold': True, 'bg_color': '#FEF9E7', 'font_color': '#7D6608',
                                  'border': 1, 'text_wrap': True})

    # ── Aba Capa ──
    ws_capa = wb.add_worksheet("Capa")
    ws_capa.set_column('A:A', 80)
    ws_capa.write('A1', 'EXPORTAÇÃO LAI — GUIAS DE TRÂNSITO VEGETAL (GTV)', fmt_titulo)
    ws_capa.write('A2', 'ADEPARÁ — Agência de Defesa Agropecuária do Estado do Pará')
    ws_capa.write('A3', 'Base legal: Lei 12.527/2011 (LAI)')
    ws_capa.write('A4', f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
    ws_capa.write('A5', f'Gerado por: {usuario.get("nome", "")} — {usuario.get("orgao", "")}')
    ws_capa.write('A6', f'Total de registros: {len(registros):,}')
    ws_capa.write('A8', 'ℹ️ GTVs não contêm CPF de pessoa física — anonimização LGPD não aplicável.', fmt_aviso)

    # ── Aba Dados ──
    ws = wb.add_worksheet("GTVs")

    cabecalhos = [label for _, label in CAMPOS_LAI_GTV]
    for col_idx, cab in enumerate(cabecalhos):
        ws.write(0, col_idx, cab, fmt_cab)
        ws.set_column(col_idx, col_idx, 22)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, 0, len(CAMPOS_LAI_GTV) - 1)

    for row_idx, reg in enumerate(registros, 1):
        fmt_linha = fmt_dado_alt if row_idx % 2 == 0 else fmt_dado
        for col_idx, (campo, _) in enumerate(CAMPOS_LAI_GTV):
            val = reg.get(campo, '') or ''
            ws.write(row_idx, col_idx, str(val) if val else '', fmt_linha)

    wb.close()
    buf_xlsx.seek(0)

    # Empacota em ZIP
    buf_zip = io.BytesIO()
    with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        nome_xlsx = f"LAI_GTV_{datetime.now().strftime('%Y%m%d')}.xlsx"
        zf.writestr(nome_xlsx, buf_xlsx.read())

    buf_zip.seek(0)
    return buf_zip
