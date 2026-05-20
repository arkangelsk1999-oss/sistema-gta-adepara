import io, os, csv, sqlite3, zipfile, json, math
from datetime import datetime
from pathlib import Path
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

STATIC  = Path(__file__).parent
DB_PATH = STATIC / "banco_gta.db"

AZUL       = "1A5276"
VERMELHO   = "FF0000"
VERM_ESC   = "8E1A0E"
BRANCO     = "FFFFFF"
CINZA      = "F2F2F2"
ROSA       = "FFF0F0"
VERDE      = "1E8449"
AMARELO    = "FFF3CD"

IBGE_UF = {
    '11': 'RO', '12': 'AC', '13': 'AM', '14': 'RR', '15': 'PA',
    '16': 'AP', '17': 'TO', '21': 'MA', '22': 'PI', '23': 'CE',
    '24': 'RN', '25': 'PB', '26': 'PE', '27': 'AL', '28': 'SE',
    '29': 'BA', '31': 'MG', '32': 'ES', '33': 'RJ', '35': 'SP',
    '41': 'PR', '42': 'SC', '43': 'RS', '50': 'MS', '51': 'MT',
    '52': 'GO', '53': 'DF',
}

OBSERVACOES_ANO = {
    2024: "ATENÇÃO: Os dados de 2024 não contêm UF de origem, Total de animais, Meio de transporte e Tipo de trânsito pois essas informações não estavam disponíveis na exportação do sistema SIGEAGRO para este período.",
    2010: "ATENÇÃO: Os dados de 2010 e 2011 são residuais (poucos registros) e não contêm UF de origem, Total de animais e Tipo de trânsito.",
    2011: "ATENÇÃO: Os dados de 2010 e 2011 são residuais (poucos registros) e não contêm UF de origem, Total de animais e Tipo de trânsito.",
}

def _converter_uf(valor):
    if not valor:
        return ''
    s = str(valor).strip()
    return IBGE_UF.get(s, s)

# ══════════════════════════════════════════════════════════════
# CAMPOS PADRONIZADOS LAI
# ══════════════════════════════════════════════════════════════
CAMPOS_LAI_PADRAO = [
    'Número da GTA',
    'Data de emissão',
    'Município de origem',
    'UF de origem',
    'Nome produtor de origem',
    'CPF/CNPJ de origem',
    'Nome estabelecimento origem',
    'Município de destino',
    'UF de destino',
    'Nome produtor de destino',
    'CPF/CNPJ destino',
    'Nome estabelecimento destino',
    'Código estabelecimento destino',
    'Finalidade do transporte',
    'Espécie',
    'Total de animais',
    'Bovinos machos 0-12 meses',
    'Bovinos fêmeas 0-12 meses',
    'Bovinos machos 13-24 meses',
    'Bovinos fêmeas 13-24 meses',
    'Bovinos machos 25-36 meses',
    'Bovinos fêmeas 25-36 meses',
    'Bovinos machos +36 meses',
    'Bovinos fêmeas +36 meses',
    'Total Bovinos Machos',
    'Total Bovinos Fêmeas',
    'Total Bovinos',
    'Meio de transporte',
    'Tipo de trânsito',
]

# ══════════════════════════════════════════════════════════════
# MAPEAMENTOS POR GRUPO DE ANOS
# ══════════════════════════════════════════════════════════════

# Grupo A: 2010-2021
MAPA_A = {
    'Número da GTA':                'Número da GTA',
    'Data de emissão':              'Data de emissão da GTA',
    'Município de origem':          'Nome do município de origem',
    'UF de origem':                 None,
    'Nome produtor de origem':      'nome do produtor de origem',
    'CPF/CNPJ de origem':           'CPF ou CNPJ do produtor de origem',
    'Nome estabelecimento origem':  'Nome estabelecimento origem',
    'Município de destino':         'nome do município destino',
    'UF de destino':                'UF destino',
    'Nome produtor de destino':     'nome do produtor de destino',
    'CPF/CNPJ destino':             'CPF ou CNPJ do produtor de destino',
    'Nome estabelecimento destino': 'nome da estabelecimento de destino',
    'Código estabelecimento destino': 'código da estabelecimento de destino',
    'Finalidade do transporte':     'Finalidade do transporte',
    'Espécie':                      'Espécie',
    'Total de animais':             None,
    'Bovinos machos 0-12 meses':    'N° de bovinos machos de 0 a 12 meses',
    'Bovinos fêmeas 0-12 meses':    'N° de bovinos fêmeas de 0 a 12 meses',
    'Bovinos machos 13-24 meses':   'N° de bovinos machos de 13 a 24 meses',
    'Bovinos fêmeas 13-24 meses':   'N° de bovinos fêmeas de 13 a 24 meses',
    'Bovinos machos 25-36 meses':   'N° de bovinos machos de 25 a 36 meses',
    'Bovinos fêmeas 25-36 meses':   'N° de bovinos fêmeas de 25 a 36 meses',
    'Bovinos machos +36 meses':     'N° de bovinos machos com mais de 36 meses',
    'Bovinos fêmeas +36 meses':     'N° de bovinos fêmeas com mais de 36 meses',
    'Total Bovinos Machos':         'Total de Bovinos Machos',
    'Total Bovinos Fêmeas':         'Total de Bovinos Fêmeas',
    'Total Bovinos':                'Total de Bovinos',
    'Meio de transporte':           'A pé, rodoviário, ferroviário, aérea, marítimo/fluvial',
    'Tipo de trânsito':             None,
}

# Grupo B: 2022-2023
MAPA_B = {
    'Número da GTA':                'Número da GTA',
    'Data de emissão':              'Data de emissão da GTA',
    'Município de origem':          'Nome do município de origem',
    'UF de origem':                 'UF ORIGEM',
    'Nome produtor de origem':      'nome do produtor de origem',
    'CPF/CNPJ de origem':           'CPF ou CNPJ do produtor de origem',
    'Nome estabelecimento origem':  'Nome estabelecimento origem',
    'Município de destino':         'nome do município destino',
    'UF de destino':                'UF destino',
    'Nome produtor de destino':     'nome do produtor de destino',
    'CPF/CNPJ destino':             'CPF ou CNPJ do produtor de destino',
    'Nome estabelecimento destino': 'nome da estabelecimento de destino',
    'Código estabelecimento destino': 'código da estabelecimento de destino',
    'Finalidade do transporte':     'Finalidade do transporte',
    'Espécie':                      'Espécie',
    'Total de animais':             'Total de animais',
    'Bovinos machos 0-12 meses':    'N° de bovinos machos de 0 a 12 meses',
    'Bovinos fêmeas 0-12 meses':    'N° de bovinos fêmeas de 0 a 12 meses',
    'Bovinos machos 13-24 meses':   'N° de bovinos machos de 13 a 24 meses',
    'Bovinos fêmeas 13-24 meses':   'N° de bovinos fêmeas de 13 a 24 meses',
    'Bovinos machos 25-36 meses':   'N° de bovinos machos de 25 a 36 meses',
    'Bovinos fêmeas 25-36 meses':   'N° de bovinos fêmeas de 25 a 36 meses',
    'Bovinos machos +36 meses':     'N° de bovinos machos com mais de 36 meses',
    'Bovinos fêmeas +36 meses':     'N° de bovinos fêmeas com mais de 36 meses',
    'Total Bovinos Machos':         'Total de Bovinos Machos',
    'Total Bovinos Fêmeas':         'Total de Bovinos Fêmeas',
    'Total Bovinos':                'Total de Bovinos',
    'Meio de transporte':           'A pé, rodoviário, ferroviário, aérea, marítimo/fluvial',
    'Tipo de trânsito':             'Tipo de Trânsito',
}

# Grupo C: 2024
MAPA_C = {
    'Número da GTA':                'gta',
    'Data de emissão':              'Data Emissão da GTA',
    'Município de origem':          'Nome Municipio Origem',
    'UF de origem':                 None,
    'Nome produtor de origem':      'Nome Produtor Origem',
    'CPF/CNPJ de origem':           'CNPJ/CPF Produtor Origem',
    'Nome estabelecimento origem':  'Nome Estabelecimento Origem',
    'Município de destino':         'Nome Municipio Destino',
    'UF de destino':                'UF Destino',
    'Nome produtor de destino':     'Nome Produtor Destino',
    'CPF/CNPJ destino':             'CNPJ/CPF Produtor Destino',
    'Nome estabelecimento destino': 'Nome Estabelecimento Destino',
    'Código estabelecimento destino': 'Codigo Estabelecimento Destino',
    'Finalidade do transporte':     'NomeFinalidade',
    'Espécie':                      'DescEspAnimal',
    'Total de animais':             None,
    'Bovinos machos 0-12 meses':    'Bovinos Machos 0 a 12',
    'Bovinos fêmeas 0-12 meses':    'Bovinos Femeas 0 a 12',
    'Bovinos machos 13-24 meses':   'Bovinos Machos 13 a 24',
    'Bovinos fêmeas 13-24 meses':   'Bovinos Femeas 13 a 24',
    'Bovinos machos 25-36 meses':   'Bovinos Machos 25 a 36',
    'Bovinos fêmeas 25-36 meses':   'Bovinos Femeas 25 a 36',
    'Bovinos machos +36 meses':     'Bovinos Machos acima 36',
    'Bovinos fêmeas +36 meses':     'Bovinos Femeas acima 36',
    'Total Bovinos Machos':         'Total Bovinos Machos',
    'Total Bovinos Fêmeas':         'Total Bovinos Femeas',
    'Total Bovinos':                'Total Bovinos',
    'Meio de transporte':           None,
    'Tipo de trânsito':             None,
}

# Grupo D: 2025
MAPA_D = {
    'Número da GTA':                'gta_numero',
    'Data de emissão':              'data_emissao',
    'Município de origem':          'origem_cidade_nome',
    'UF de origem':                 'origem_estado_nome',
    'Nome produtor de origem':      'origem_nome',
    'CPF/CNPJ de origem':           'origem_identificacao',
    'Nome estabelecimento origem':  'origem_estabelecimento',
    'Município de destino':         'destinatario_cidade_nome',
    'UF de destino':                'destinatario_estado_nome',
    'Nome produtor de destino':     'destinatario_nome',
    'CPF/CNPJ destino':             'destinatario_identificacao',
    'Nome estabelecimento destino': 'destinatario_estabelecimento',
    'Código estabelecimento destino': 'destinatario_codigo_estabelecimento',
    'Finalidade do transporte':     'finalidade',
    'Espécie':                      'taxonomia',
    'Total de animais':             'total_animais',
    'Bovinos machos 0-12 meses':    'BOVINO, MACHO, 0 A 12 MESES',
    'Bovinos fêmeas 0-12 meses':    'BOVINO, FÊMEA, 0 A 12 MESES',
    'Bovinos machos 13-24 meses':   'BOVINO, MACHO, 13 A 24 MESES',
    'Bovinos fêmeas 13-24 meses':   'BOVINO, FÊMEA, 13 A 24 MESES',
    'Bovinos machos 25-36 meses':   'BOVINO, MACHO, 25 A 36 MESES',
    'Bovinos fêmeas 25-36 meses':   'BOVINO, FÊMEA, 25 A 36 MESES',
    'Bovinos machos +36 meses':     'BOVINO, MACHO, ACIMA DE 36 MESES',
    'Bovinos fêmeas +36 meses':     'BOVINO, FÊMEA, ACIMA DE 36 MESES',
    'Total Bovinos Machos':         None,
    'Total Bovinos Fêmeas':         None,
    'Total Bovinos':                None,
    'Meio de transporte':           None,
    'Tipo de trânsito':             None,
}

# Colunas CPF de origem e destino por grupo
CPF_ORIGEM_POR_GRUPO = {
    'A': 'CPF ou CNPJ do produtor de origem',
    'B': 'CPF ou CNPJ do produtor de origem',
    'C': 'CNPJ/CPF Produtor Origem',
    'D': 'origem_identificacao',
}

CPF_DESTINO_POR_GRUPO = {
    'A': 'CPF ou CNPJ do produtor de destino',
    'B': 'CPF ou CNPJ do produtor de destino',
    'C': 'CNPJ/CPF Produtor Destino',
    'D': 'destinatario_identificacao',
}

UF_PRECISA_CONVERTER = {
    'C': {'UF de destino'},
}

def _get_grupo(ano):
    if ano <= 2021:   return 'A'
    elif ano <= 2023: return 'B'
    elif ano == 2024: return 'C'
    else:             return 'D'

def _get_mapa(ano):
    return {'A': MAPA_A, 'B': MAPA_B, 'C': MAPA_C, 'D': MAPA_D}[_get_grupo(ano)]

def _extrair_valor_padrao(dados, campo_padrao, ano):
    mapa     = _get_mapa(ano)
    grupo    = _get_grupo(ano)
    col_real = mapa.get(campo_padrao)
    if col_real is None:
        return ''
    val = dados.get(col_real, '')
    if val is None:
        return ''
    if campo_padrao in UF_PRECISA_CONVERTER.get(grupo, set()):
        val = _converter_uf(str(val))
    return val

def _registro_valido(dados, ano):
    mapa    = _get_mapa(ano)
    col_gta = mapa.get('Número da GTA')
    if not col_gta:
        return True
    val = dados.get(col_gta)
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return False
    return str(val).strip() != ''

def _limpar_nan(dados):
    return {k: (None if isinstance(v, float) and math.isnan(v) else v) for k, v in dados.items()}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size)

def _align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def _periodo_banco():
    try:
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = None
        row = conn.execute("SELECT MAX(ano) FROM gtas").fetchone()
        conn.close()
        if row and row[0]:
            return 2012, int(row[0])
    except:
        pass
    return 2012, 2025

def _anonimizar_cpf_cnpj(valor):
    """CPF (11 dígitos) → anonimizado. CNPJ (14 dígitos) → mantém."""
    if not valor:
        return ''
    digitos = ''.join(filter(str.isdigit, str(valor)))
    if len(digitos) == 14:
        return digitos
    return '***ANONIMIZADO***'

def _is_cpf(valor):
    """Retorna True se o valor for CPF (11 dígitos)."""
    if not valor:
        return False
    digitos = ''.join(filter(str.isdigit, str(valor)))
    return len(digitos) == 11

def _anonimizar_nome_se_cpf(nome, cpf_cnpj):
    """Anonimiza o nome se o CPF/CNPJ for CPF (pessoa física)."""
    if _is_cpf(cpf_cnpj):
        return '***ANONIMIZADO***'
    return nome or ''


# ── Excel resultado busca ─────────────────────────────────────
def gerar_excel_resultado(resultado, nome_busca, cpf_busca, usuario=None):
    wb = Workbook()
    wb.remove(wb.active)
    anos = sorted(resultado.keys())

    ano_min, ano_max = _periodo_banco()
    periodo_banco = f"{ano_min} a {ano_max}" if ano_min else "N/D"

    usuario_nome = usuario.get('nome', '') if usuario else ''
    usuario_cpf  = usuario.get('cpf', '')  if usuario else ''
    agora        = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    wb.properties.creator        = usuario_nome
    wb.properties.lastModifiedBy = usuario_nome
    wb.properties.description    = f"Gerado por: {usuario_nome} | CPF: {usuario_cpf} | Data: {agora}"
    wb.properties.subject        = "Sistema de Consulta a Histórico de Emissões de GTAs — ADEPARÁ"
    wb.properties.keywords       = f"{usuario_nome} {usuario_cpf} {agora}"

    ws = wb.create_sheet("RESUMO", 0)
    ws.merge_cells("A1:D1")
    ws["A1"] = "AGÊNCIA DE DEFESA AGROPECUÁRIA DO ESTADO DO PARÁ"
    ws["A1"].font      = Font(bold=True, size=13, color=AZUL)
    ws["A1"].alignment = _align()

    ws["A3"] = "Produtor pesquisado:"
    ws["B3"] = (nome_busca or '').upper()
    ws["A4"] = "CPF/CNPJ:"
    ws["B4"] = cpf_busca or ''
    ws["A5"] = "Data do relatório:"
    ws["B5"] = agora
    ws["A6"] = "Gerado por:"
    ws["B6"] = f"{usuario_nome} — CPF: {usuario_cpf}"
    ws["A7"] = "Base de dados:"
    ws["B7"] = f"GTAs emitidas no Pará — {periodo_banco}"
    ws["A7"].font = Font(bold=True, color=AZUL)
    ws["B7"].font = Font(bold=True, color=AZUL)

    for col, titulo in enumerate(["ANO","GTAs como REMETENTE","GTAs como DESTINATÁRIO","TOTAL"], 1):
        c = ws.cell(row=9, column=col, value=titulo)
        c.fill = _fill(AZUL); c.font = _font(bold=True); c.alignment = _align()

    tot_o = tot_d = 0
    for i, ano in enumerate(anos):
        qo = len(resultado[ano]['origem'])
        qd = len(resultado[ano]['destino'])
        ws.cell(row=10+i, column=1, value=str(ano))
        ws.cell(row=10+i, column=2, value=qo)
        ws.cell(row=10+i, column=3, value=qd)
        ws.cell(row=10+i, column=4, value=qo+qd)
        tot_o += qo; tot_d += qd

    lr = 10 + len(anos)
    for col, val in enumerate(["TOTAL", tot_o, tot_d, tot_o+tot_d], 1):
        c = ws.cell(row=lr, column=col, value=val)
        c.font = Font(bold=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12

    COLS_INTERNAS = {'_ano','_arquivo','_formato'}
    for ano in anos:
        orig = resultado[ano]['origem']
        dest = resultado[ano]['destino']
        cols = resultado[ano]['colunas']
        cols = [c for c in cols if c not in COLS_INTERNAS]
        if not orig and not dest:
            continue

        ws_ano = wb.create_sheet(str(ano))
        linha  = 1

        def escrever_bloco(dados, fill_header, fill_zebra):
            nonlocal linha
            if not dados:
                return
            for ci, col in enumerate(cols, 1):
                c = ws_ano.cell(row=linha, column=ci, value=col)
                c.fill = _fill(fill_header)
                c.font = _font(bold=True)
                c.alignment = _align()
            ws_ano.row_dimensions[linha].height = 28
            linha += 1
            for ri, row_dict in enumerate(dados):
                for ci, col in enumerate(cols, 1):
                    val = row_dict.get(col, '')
                    ws_ano.cell(row=linha, column=ci, value=val)
                    if ri % 2 == 0:
                        ws_ano.cell(row=linha, column=ci).fill = _fill(fill_zebra)
                linha += 1

        escrever_bloco(orig, AZUL, CINZA)

        for ci in range(1, len(cols)+1):
            c = ws_ano.cell(row=linha, column=ci)
            c.fill = _fill(VERMELHO)
            c.font = _font(bold=True)
            c.alignment = _align()
        ws_ano.cell(row=linha, column=1).value = "▼  DESTINATÁRIO  ▼"
        ws_ano.row_dimensions[linha].height = 18
        linha += 1

        escrever_bloco(dest, VERM_ESC, "FFF0F0")

        for ci, col in enumerate(cols, 1):
            letra = get_column_letter(ci)
            tam = min(max(len(str(col)), 10) + 2, 40)
            ws_ano.column_dimensions[letra].width = tam

        ultima_col = get_column_letter(len(cols))
        ws_ano.auto_filter.ref = f"A1:{ultima_col}1"
        ws_ano.freeze_panes = "A2"

    ws_r = wb.create_sheet("_rastreio")
    ws_r["A1"] = "Gerado por"
    ws_r["B1"] = usuario_nome
    ws_r["A2"] = "CPF"
    ws_r["B2"] = usuario_cpf
    ws_r["A3"] = "Data/Hora"
    ws_r["B3"] = agora
    ws_r["A4"] = "Sistema"
    ws_r["B4"] = "Sistema de Consulta a Histórico de Emissões de GTAs — ADEPARÁ"
    ws_r["A5"] = "Produtor pesquisado"
    ws_r["B5"] = (nome_busca or '').upper()
    ws_r["A6"] = "CPF/CNPJ pesquisado"
    ws_r["B6"] = cpf_busca or ''
    ws_r["A7"] = "Base de dados"
    ws_r["B7"] = f"GTAs emitidas no Pará — {periodo_banco}"

    senha_protecao = ''.join(filter(str.isdigit, usuario_cpf)) if usuario_cpf else 'adepara2025'
    ws_r.protection.sheet    = True
    ws_r.protection.password = senha_protecao
    ws_r.sheet_state = 'hidden'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── CSV (somente founder) ─────────────────────────────────────
def gerar_csv_resultado(resultado, nome_busca, cpf_busca, usuario=None):
    COLS_INTERNAS = {'_ano','_arquivo','_formato'}
    anos = sorted(resultado.keys())

    buf = io.StringIO()
    writer = None

    for ano in anos:
        orig = resultado[ano]['origem']
        dest = resultado[ano]['destino']
        cols = [c for c in resultado[ano]['colunas'] if c not in COLS_INTERNAS]

        todos = []
        for row in orig:
            r = {c: row.get(c, '') for c in cols}
            r['__PAPEL__'] = 'REMETENTE'
            r['__ANO__']   = str(ano)
            todos.append(r)
        for row in dest:
            r = {c: row.get(c, '') for c in cols}
            r['__PAPEL__'] = 'DESTINATARIO'
            r['__ANO__']   = str(ano)
            todos.append(r)

        if not todos:
            continue

        fieldnames = ['__ANO__', '__PAPEL__'] + cols

        if writer is None:
            writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=';',
                                    extrasaction='ignore')
            writer.writeheader()

        writer.writerows(todos)

    buf.seek(0)
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


# ══════════════════════════════════════════════════════════════
# EXCEL LAI — ZIP com um arquivo por ano (xlsxwriter)
# ══════════════════════════════════════════════════════════════
def _gerar_excel_lai_ano_xlsx(ano, rows_ano, usuario_nome, usuario_cpf, agora, periodo_banco, ano_ini, ano_fim):
    """Gera Excel para um único ano do LAI usando xlsxwriter."""

    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf, {'in_memory': True, 'constant_memory': True})

    fmt_titulo = wb.add_format({'bold': True, 'font_size': 13, 'font_color': '#1A5276', 'align': 'center'})
    fmt_label  = wb.add_format({'bold': True})
    fmt_obs    = wb.add_format({'font_color': '#8B4513', 'bg_color': '#FFF3CD', 'text_wrap': True})
    fmt_header = wb.add_format({
        'bold': True, 'font_color': 'white', 'bg_color': '#1A5276',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1
    })
    fmt_normal = wb.add_format({'border': 1})
    fmt_zebra  = wb.add_format({'border': 1, 'bg_color': '#F2F2F2'})

    # ── CAPA ──────────────────────────────────────────────────
    ws_capa = wb.add_worksheet('CAPA')
    ws_capa.set_column('A:A', 25)
    ws_capa.set_column('B:E', 30)
    ws_capa.merge_range('A1:E1', 'AGÊNCIA DE DEFESA AGROPECUÁRIA DO ESTADO DO PARÁ', fmt_titulo)

    infos = [
        ('Base de dados:',      f'GTAs emitidas no Pará — {periodo_banco}'),
        ('Período solicitado:', f'{ano_ini} a {ano_fim}'),
        ('Ano deste arquivo:',  str(ano)),
        ('Data do relatório:',  agora),
        ('Gerado por:',         f'{usuario_nome} — CPF: {usuario_cpf}'),
        ('Anonimização:',       'Dados de pessoas físicas anonimizados conforme LGPD. Dados de pessoas jurídicas (CNPJ) mantidos por serem públicos.'),
        ('Base legal:',         'Lei nº 12.527/2011 (LAI) | Lei nº 13.709/2018 (LGPD)'),
    ]
    for i, (label, valor) in enumerate(infos, 3):
        ws_capa.write(i, 0, label, fmt_label)
        ws_capa.write(i, 1, valor)

    obs = OBSERVACOES_ANO.get(ano)
    linha_obs = len(infos) + 4
    if obs:
        ws_capa.write(linha_obs, 0, '⚠️ Observação:', fmt_label)
        ws_capa.merge_range(linha_obs, 1, linha_obs, 4, obs, fmt_obs)
        ws_capa.set_row(linha_obs, 40)
        linha_total = linha_obs + 2
    else:
        linha_total = linha_obs

    # ── DADOS ─────────────────────────────────────────────────
    ws_dados = wb.add_worksheet(f'GTAs_{ano}')

    for ci, col in enumerate(CAMPOS_LAI_PADRAO):
        ws_dados.write(0, ci, col, fmt_header)
        ws_dados.set_column(ci, ci, max(len(col) + 2, 14))
    ws_dados.set_row(0, 28)
    ws_dados.freeze_panes(1, 0)
    ws_dados.autofilter(0, 0, 0, len(CAMPOS_LAI_PADRAO) - 1)

    grupo = _get_grupo(ano)
    total = 0

    for row in rows_ano:
        dados = _limpar_nan(json.loads(row['dados_json']))

        if not _registro_valido(dados, ano):
            continue

        # Busca CPF/CNPJ de origem e destino para aplicar regras
        col_cpf_orig = CPF_ORIGEM_POR_GRUPO[grupo]
        col_cpf_dest = CPF_DESTINO_POR_GRUPO[grupo]
        cpf_cnpj_orig = dados.get(col_cpf_orig, '') or ''
        cpf_cnpj_dest = dados.get(col_cpf_dest, '') or ''

        fmt_linha = fmt_zebra if total % 2 == 0 else fmt_normal

        for ci, campo_padrao in enumerate(CAMPOS_LAI_PADRAO):
            val = _extrair_valor_padrao(dados, campo_padrao, ano)

            # ── Regras de anonimização ────────────────────────
            if campo_padrao == 'CPF/CNPJ de origem':
                val = _anonimizar_cpf_cnpj(cpf_cnpj_orig)

            elif campo_padrao == 'Nome produtor de origem':
                val = _anonimizar_nome_se_cpf(val, cpf_cnpj_orig)

            elif campo_padrao == 'Nome estabelecimento origem':
                val = _anonimizar_nome_se_cpf(val, cpf_cnpj_orig)

            elif campo_padrao == 'CPF/CNPJ destino':
                val = _anonimizar_cpf_cnpj(cpf_cnpj_dest)

            elif campo_padrao == 'Nome produtor de destino':
                val = _anonimizar_nome_se_cpf(val, cpf_cnpj_dest)

            elif campo_padrao == 'Nome estabelecimento destino':
                val = _anonimizar_nome_se_cpf(val, cpf_cnpj_dest)

            ws_dados.write(total + 1, ci, str(val) if val else '', fmt_linha)

        total += 1

    ws_capa.write(linha_total, 0, 'Total de registros:', fmt_label)
    ws_capa.write(linha_total, 1, total)

    wb.close()
    buf.seek(0)
    return buf, total


def gerar_excel_lai(resultado, ano_ini, ano_fim, usuario=None, callback_progresso=None):
    """Gera ZIP com um Excel por ano para atender pedidos LAI. Usa xlsxwriter."""
    usuario_nome  = usuario.get('nome', '') if usuario else ''
    usuario_cpf   = usuario.get('cpf', '')  if usuario else ''
    agora         = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ano_min, ano_max = _periodo_banco()
    periodo_banco = f"{ano_min} a {ano_max}" if ano_min else 'N/D'

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    anos = [row[0] for row in conn.execute(
        "SELECT DISTINCT ano FROM gtas WHERE ano >= ? AND ano <= ? ORDER BY ano",
        (int(ano_ini), int(ano_fim))
    ).fetchall()]

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for idx, ano in enumerate(anos):
            rows_ano = conn.execute(
                "SELECT dados_json FROM gtas WHERE ano = ? ORDER BY rowid",
                (ano,)
            ).fetchall()

            if not rows_ano:
                continue

            excel_buf, total = _gerar_excel_lai_ano_xlsx(
                ano, rows_ano, usuario_nome, usuario_cpf,
                agora, periodo_banco, ano_ini, ano_fim
            )

            zf.writestr(f'LAI_GTA_ADEPARA_{ano}.xlsx', excel_buf.read())

            if callback_progresso:
                callback_progresso(ano, idx + 1, len(anos), total)

            del rows_ano
            del excel_buf

    conn.close()
    zip_buf.seek(0)
    return zip_buf


# ── PDF de Auditoria ──────────────────────────────────────────
def gerar_pdf_auditoria(rows, gerado_por, data_ini='', data_fim=''):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=4*cm, bottomMargin=1.5*cm
    )

    cor_azul = colors.HexColor('#1A5276')

    estilo_celula = ParagraphStyle(
        'celula', fontSize=6.5, leading=8, wordWrap='CJK',
    )
    estilo_header = ParagraphStyle(
        'header', fontSize=7, leading=9, textColor=colors.white,
        fontName='Helvetica-Bold', alignment=1,
    )

    elementos = []

    periodo = f"{data_ini or 'início'} a {data_fim or 'hoje'}"
    elementos.append(Paragraph(
        "<b>Relatório de Auditoria de Acessos</b>",
        ParagraphStyle('titulo', fontSize=13, textColor=cor_azul,
                       alignment=TA_CENTER, spaceAfter=4)
    ))
    elementos.append(Paragraph(
        f"Período: {periodo} &nbsp;&nbsp; | &nbsp;&nbsp; Gerado por: {gerado_por.get('nome','')} &nbsp;&nbsp; | &nbsp;&nbsp; {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle('sub', fontSize=9, textColor=colors.grey,
                       alignment=TA_CENTER, spaceAfter=12)
    ))

    cabecalho = [
        Paragraph("Data/Hora", estilo_header),
        Paragraph("Usuário", estilo_header),
        Paragraph("Órgão", estilo_header),
        Paragraph("IP", estilo_header),
        Paragraph("Localidade", estilo_header),
        Paragraph("CPF/CNPJ pesquisado", estilo_header),
        Paragraph("Nome pesquisado", estilo_header),
    ]

    dados_tabela = [cabecalho]
    for r in rows:
        dados_tabela.append([
            Paragraph(r.get('data_hora',''), estilo_celula),
            Paragraph(r.get('usuario_nome',''), estilo_celula),
            Paragraph(r.get('orgao',''), estilo_celula),
            Paragraph(r.get('ip',''), estilo_celula),
            Paragraph(r.get('localidade',''), estilo_celula),
            Paragraph(r.get('cpf_cnpj_pesquisado',''), estilo_celula),
            Paragraph(r.get('nome_pesquisado',''), estilo_celula),
        ])

    col_widths = [2.8*cm, 3.2*cm, 2.5*cm, 2.2*cm, 2.0*cm, 2.5*cm, 4.8*cm]

    tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), cor_azul),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 3),
        ('RIGHTPADDING',  (0,0), (-1,-1), 3),
    ]))
    elementos.append(tabela)

    brasao_path  = str(STATIC / 'brasao_para.png')
    adepara_path = str(STATIC / 'adepara_logo.png')

    def header(canvas, doc):
        canvas.saveState()
        w, h = A4

        if os.path.exists(brasao_path):
            canvas.drawImage(brasao_path, 1.5*cm, h-3.5*cm, width=2*cm, height=2.5*cm,
                             preserveAspectRatio=True, mask='auto')

        if os.path.exists(adepara_path):
            canvas.drawImage(adepara_path, w-4.5*cm, h-3.5*cm, width=3*cm, height=2*cm,
                             preserveAspectRatio=True, mask='auto')

        canvas.setFont("Helvetica-Bold", 10)
        canvas.setFillColor(cor_azul)
        canvas.drawCentredString(w/2, h-1.5*cm, "AGÊNCIA DE DEFESA AGROPECUÁRIA DO ESTADO DO PARÁ")

        canvas.setStrokeColor(cor_azul)
        canvas.setLineWidth(0.5)
        canvas.line(1.5*cm, h-3.7*cm, w-1.5*cm, h-3.7*cm)

        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(w/2, 0.8*cm,
            f"Página {doc.page} — Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Consulta GTA")
        canvas.restoreState()

    doc.build(elementos, onFirstPage=header, onLaterPages=header)
    buf.seek(0)
    return buf