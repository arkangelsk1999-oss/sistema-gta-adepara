"""
relatorio_gta2026.py
Geração de Excel, CSV e LAI para o módulo GTA 2026.
"""

import io
import re
from datetime import datetime

# ─────────────────────────────────────────────
# CAMPOS PARA EXPORTAÇÃO NORMAL
# ─────────────────────────────────────────────

CAMPOS_EXIBICAO = [
    ("gta_numero",                    "Nº GTA"),
    ("data_emissao",                  "Data Emissão"),
    ("finalidade",                    "Finalidade"),
    ("situacao_gta",                  "Situação"),
    ("origem_nome",                   "Produtor Origem"),
    ("origem_identificacao",          "CPF/CNPJ Origem"),
    ("origem_estabelecimento",        "Estabelecimento Origem"),
    ("origem_cidade_nome",            "Município Origem"),
    ("origem_estado_nome",            "UF Origem"),
    ("destinatario_nome",             "Produtor Destino"),
    ("destinatario_identificacao",    "CPF/CNPJ Destino"),
    ("destinatario_estabelecimento",  "Estabelecimento Destino"),
    ("destinatario_cidade_nome",      "Município Destino"),
    ("destinatario_estado_nome",      "UF Destino"),
    ("total_animais",                 "Total Animais"),
    ("taxonomia",                     "Espécie"),
    ("transporte",                    "Transporte"),
    ("emitida_por_nome",              "Emitido Por"),
    ("valor_dae",                     "Valor DAE"),
    ("valor_fundepec",                "Valor FUNDEPEC"),
    ("valor_total",                   "Valor Total"),
]

# ─────────────────────────────────────────────
# CAMPOS LAI PADRONIZADOS
# ─────────────────────────────────────────────

CAMPOS_LAI_2026 = [
    ("gta_numero",                        "Número da GTA"),
    ("data_emissao",                      "Data de Emissão"),
    ("origem_cidade_nome",                "Município de Origem"),
    ("origem_estado_nome",                "UF de Origem"),
    ("origem_nome",                       "Nome Produtor Origem"),
    ("origem_identificacao",              "CPF/CNPJ Origem"),
    ("origem_estabelecimento",            "Nome Estabelecimento Origem"),
    ("destinatario_cidade_nome",          "Município de Destino"),
    ("destinatario_estado_nome",          "UF de Destino"),
    ("destinatario_nome",                 "Nome Produtor Destino"),
    ("destinatario_identificacao",        "CPF/CNPJ Destino"),
    ("destinatario_estabelecimento",      "Nome Estabelecimento Destino"),
    ("destinatario_codigo_estabelecimento", "Código Estabelecimento Destino"),
    ("finalidade",                        "Finalidade do Transporte"),
    ("taxonomia",                         "Espécie"),
    ("total_animais",                     "Total de Animais"),
    ("bovino_macho_0_12",                 "Bovinos Machos 0-12 Meses"),
    ("bovino_femea_0_12",                 "Bovinos Fêmeas 0-12 Meses"),
    ("bovino_macho_13_24",                "Bovinos Machos 13-24 Meses"),
    ("bovino_femea_13_24",                "Bovinos Fêmeas 13-24 Meses"),
    ("bovino_macho_25_36",                "Bovinos Machos 25-36 Meses"),
    ("bovino_femea_25_36",                "Bovinos Fêmeas 25-36 Meses"),
    ("bovino_macho_acima_36",             "Bovinos Machos +36 Meses"),
    ("bovino_femea_acima_36",             "Bovinos Fêmeas +36 Meses"),
    ("transporte",                        "Meio de Transporte"),
]


# ─────────────────────────────────────────────
# ANONIMIZAÇÃO LGPD
# ─────────────────────────────────────────────

def _is_cpf(val):
    if not val:
        return False
    s = re.sub(r'[^\d]', '', str(val))
    return len(s) == 11


def _anonimizar_se_cpf(val):
    return '***ANONIMIZADO***' if _is_cpf(val) else val


def _anonimizar_nome_se_cpf(nome, doc):
    return '***ANONIMIZADO***' if _is_cpf(doc) else nome


# ─────────────────────────────────────────────
# EXCEL — BUSCA NORMAL
# ─────────────────────────────────────────────

def gerar_excel_2026(resultado, termo_busca, cpf_busca, usuario):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    azul_escuro = PatternFill("solid", fgColor="1A5276")
    azul_claro  = PatternFill("solid", fgColor="D6EAF8")
    verde       = PatternFill("solid", fgColor="1E8449")
    branco_font = Font(color="FFFFFF", bold=True, size=11)
    bold        = Font(bold=True)

    dados_origem  = resultado.get('2026', {}).get('origem', [])
    dados_destino = resultado.get('2026', {}).get('destino', [])

    for titulo_aba, dados in [("Como Origem", dados_origem), ("Como Destino", dados_destino)]:
        if not dados:
            continue
        ws = wb.create_sheet(titulo_aba)

        # Cabeçalho
        cabecalhos = [label for _, label in CAMPOS_EXIBICAO]
        for col_idx, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_idx, value=cab)
            cell.fill = azul_escuro
            cell.font = branco_font
            cell.alignment = Alignment(horizontal='center')

        # Dados
        for row_idx, reg in enumerate(dados, 2):
            fill = azul_claro if row_idx % 2 == 0 else PatternFill()
            for col_idx, (campo, _) in enumerate(CAMPOS_EXIBICAO, 1):
                val = reg.get(campo, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
                cell.fill = fill

        # Ajusta largura
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Aba rastreio oculta
    ws_r = wb.create_sheet("_rastreio")
    ws_r.sheet_state = 'hidden'
    cpf_usuario = usuario.get('cpf', '00000000000')
    ws_r['A1'] = "Sistema GTA ADEPARÁ — Módulo 2026"
    ws_r['A2'] = f"Gerado por: {usuario.get('nome', '')}"
    ws_r['A3'] = f"E-mail: {usuario.get('email', '')}"
    ws_r['A4'] = f"Órgão: {usuario.get('orgao', '')}"
    ws_r['A5'] = f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_r['A6'] = f"Termo buscado: {termo_busca}"
    ws_r['A7'] = f"CPF/CNPJ: {cpf_busca}"
    ws_r.protection.sheet = True
    ws_r.protection.password = cpf_usuario

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# CSV — BUSCA NORMAL
# ─────────────────────────────────────────────

def gerar_csv_2026(resultado, termo_busca, cpf_busca, usuario):
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')

    cabecalhos = [label for _, label in CAMPOS_EXIBICAO]
    writer.writerow(cabecalhos)

    for tipo in ['origem', 'destino']:
        for reg in resultado.get('2026', {}).get(tipo, []):
            linha = [reg.get(campo, '') or '' for campo, _ in CAMPOS_EXIBICAO]
            writer.writerow(linha)

    buf.seek(0)
    return io.BytesIO(buf.getvalue().encode('utf-8-sig'))


# ─────────────────────────────────────────────
# LAI 2026 — Excel com anonimização LGPD
# ─────────────────────────────────────────────

def gerar_excel_lai_2026(registros, usuario):
    """
    Gera Excel LAI para todos os registros de 2026.
    Anonimização LGPD: CPF → ***ANONIMIZADO***, CNPJ → mantido.
    Retorna BytesIO com o arquivo ZIP contendo o Excel.
    """
    try:
        import xlsxwriter
        import zipfile
    except ImportError:
        raise ImportError("xlsxwriter não instalado. Execute: pip install xlsxwriter")

    buf_xlsx = io.BytesIO()
    wb = xlsxwriter.Workbook(buf_xlsx, {'in_memory': True})

    # Formatos
    fmt_titulo  = wb.add_format({'bold': True, 'font_size': 13, 'font_color': '#1A5276'})
    fmt_cab     = wb.add_format({'bold': True, 'bg_color': '#1A5276', 'font_color': '#FFFFFF',
                                  'border': 1, 'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
    fmt_dado    = wb.add_format({'border': 1, 'text_wrap': False})
    fmt_dado_alt= wb.add_format({'border': 1, 'bg_color': '#EBF5FB', 'text_wrap': False})
    fmt_anon    = wb.add_format({'border': 1, 'font_color': '#C0392B', 'italic': True})
    fmt_aviso   = wb.add_format({'bold': True, 'bg_color': '#FEF9E7', 'font_color': '#7D6608',
                                  'border': 1, 'text_wrap': True})

    # ── Aba Capa ──
    ws_capa = wb.add_worksheet("Capa")
    ws_capa.set_column('A:A', 80)
    ws_capa.write('A1', 'EXPORTAÇÃO LAI — GUIAS DE TRÂNSITO ANIMAL (GTA) 2026', fmt_titulo)
    ws_capa.write('A2', f'ADEPARÁ — Agência de Defesa Agropecuária do Estado do Pará')
    ws_capa.write('A3', f'Base legal: Lei 12.527/2011 (LAI) + Lei 13.709/2018 (LGPD)')
    ws_capa.write('A4', f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
    ws_capa.write('A5', f'Gerado por: {usuario.get("nome", "")} — {usuario.get("orgao", "")}')
    ws_capa.write('A6', f'Total de registros: {len(registros):,}')
    ws_capa.write('A8', '⚠️ LGPD: CPF (pessoa física) anonimizado. CNPJ (pessoa jurídica) mantido por ser dado público.', fmt_aviso)

    # ── Aba Dados ──
    ws = wb.add_worksheet("GTAs 2026")

    cabecalhos = [label for _, label in CAMPOS_LAI_2026]
    for col_idx, cab in enumerate(cabecalhos):
        ws.write(0, col_idx, cab, fmt_cab)
        ws.set_column(col_idx, col_idx, 20)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, 0, len(CAMPOS_LAI_2026) - 1)

    for row_idx, reg in enumerate(registros, 1):
        fmt_linha = fmt_dado_alt if row_idx % 2 == 0 else fmt_dado

        for col_idx, (campo, _) in enumerate(CAMPOS_LAI_2026):
            val = reg.get(campo, '') or ''

            # Anonimização LGPD
            if campo == 'origem_identificacao':
                val = _anonimizar_se_cpf(val)
            elif campo == 'destinatario_identificacao':
                val = _anonimizar_se_cpf(val)
            elif campo == 'origem_nome':
                val = _anonimizar_nome_se_cpf(val, reg.get('origem_identificacao', ''))
            elif campo == 'destinatario_nome':
                val = _anonimizar_nome_se_cpf(val, reg.get('destinatario_identificacao', ''))
            elif campo == 'origem_estabelecimento':
                val = _anonimizar_nome_se_cpf(val, reg.get('origem_identificacao', ''))
            elif campo == 'destinatario_estabelecimento':
                val = _anonimizar_nome_se_cpf(val, reg.get('destinatario_identificacao', ''))

            fmt_usar = fmt_anon if val == '***ANONIMIZADO***' else fmt_linha
            ws.write(row_idx, col_idx, str(val) if val else '', fmt_usar)

    wb.close()
    buf_xlsx.seek(0)

    # Empacota em ZIP
    buf_zip = io.BytesIO()
    with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        nome_xlsx = f"LAI_GTA_2026_{datetime.now().strftime('%Y%m%d')}.xlsx"
        zf.writestr(nome_xlsx, buf_xlsx.read())

    buf_zip.seek(0)
    return buf_zip
